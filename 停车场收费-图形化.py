import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import threading
import os
import datetime

# ==========================================
# 核心逻辑类 (基于 V7 修改，增加回调支持)
# ==========================================
class ParkingLogic:
    # def run(self, hotel_file, total_file, output_file, base_quota, boss_plates_str, log_callback, progress_callback):
    # 【修改为】: 增加 floating_quota 参数
    def run(self, hotel_file, total_file, output_file, base_quota, floating_quota, boss_plates_str, log_callback,
            progress_callback):
        try:
            log_callback("正在初始化数据...")
            
            # 1. 基础配置
            boss_plates = set([p.strip().upper() for p in boss_plates_str.replace('，', ',').split(',') if p.strip()])
            
            # 浮动车位数量 = 老总车牌数量
            # floating_quota = len(boss_plates)
            total_capacity = base_quota + floating_quota
            
            log_callback(f"【规则核对】")
            log_callback(f"  - 基础免费车位: {base_quota}")
            log_callback(f"  - 浮动车位(老总): {floating_quota}")
            log_callback(f"  - 总免费池容量: {total_capacity}")
            
            # 读取文件
            try:
                hotel_df = pd.read_excel(hotel_file)
                total_df = pd.read_excel(total_file)
            except Exception as e:
                raise Exception(f"Excel读取失败: {str(e)}")
            
            # 清洗列名
            hotel_df.columns = [c.strip() for c in hotel_df.columns]
            total_df.columns = [c.strip() for c in total_df.columns]
            if '车牌号/卡号' in total_df.columns:
                total_df.rename(columns={'车牌号/卡号': '车牌号码'}, inplace=True)
                
            # 时间转换
            for df in [hotel_df, total_df]:
                df['入场时间'] = pd.to_datetime(df['入场时间'])
                df['出场时间'] = pd.to_datetime(df['出场时间'])

            # 身份标记
            total_df['车辆类型'] = '其他'
            total_df.loc[total_df['车牌号码'].isin(boss_plates), '车辆类型'] = '老总'
            
            hotel_keys = set(zip(hotel_df['车牌号码'], hotel_df['入场时间']))
            def identify(row):
                if row['车辆类型'] == '老总': return '老总'
                if (row['车牌号码'], row['入场时间']) in hotel_keys: return '酒楼客'
                return '其他'
            
            total_df['车辆类型'] = total_df.apply(identify, axis=1)
            target_df = total_df[total_df['车辆类型'].isin(['酒楼客', '老总'])].copy()
            target_df.drop_duplicates(subset=['车牌号码', '入场时间', '出场时间'], inplace=True)

            # ==========================
            # 事件推演 (V7算法)
            # ==========================
            all_time_points = set()
            all_time_points.update(target_df['入场时间'])
            all_time_points.update(target_df['出场时间'])
            timeline = sorted(list(all_time_points))
            
            # 初始化数据
            grid_data = {(row['车牌号码'], row['入场时间']): {} for _, row in target_df.iterrows()}
            permanently_charged_records = set()
            target_df['判定结果'] = '免费'
            
            log_callback(f"开始时间轴推演，共 {len(timeline)} 个时间点...")
            
            total_steps = len(timeline)
            
            for i in range(len(timeline) - 1):
                # 更新进度条 (每100步更新一次，避免界面卡顿)
                if i % 50 == 0:
                    progress = (i / total_steps) * 100
                    progress_callback(progress)
                
                current_time = timeline[i]
                
                # 1. 找出此刻在场车辆
                active_cars = target_df[
                    (target_df['入场时间'] <= current_time) & 
                    (target_df['出场时间'] > current_time)
                ]
                
                if active_cars.empty:
                    continue
                    
                active_bosses = active_cars[active_cars['车辆类型'] == '老总']
                active_guests = active_cars[active_cars['车辆类型'] == '酒楼客'].sort_values(by='入场时间')
                
                # 2. 计算配额
                current_boss_count = len(active_bosses)
                available_free_slots = total_capacity - current_boss_count
                
                # 记录老总
                for _, row in active_bosses.iterrows():
                    grid_data[(row['车牌号码'], row['入场时间'])][current_time] = '免'

                # 3. 分配名额 (V7: 付费不占名额)
                used_free_slots_count = 0
                
                for idx, row in active_guests.iterrows():
                    plate = row['车牌号码']
                    entry_time = row['入场时间']
                    record_key = (plate, entry_time)
                    
                    is_already_sticky = record_key in permanently_charged_records
                    
                    if is_already_sticky:
                        status = '收'
                    else:
                        if used_free_slots_count < available_free_slots:
                            status = '免'
                            used_free_slots_count += 1
                        else:
                            status = '收'
                            permanently_charged_records.add(record_key)
                            target_df.at[idx, '判定结果'] = '需要收费'
                    
                    grid_data[(plate, entry_time)][current_time] = status

            # 进度完成
            progress_callback(90)
            log_callback("计算完成，正在生成Excel文件...")

            # ==========================
            # 导出 Excel
            # ==========================
            target_df.loc[target_df['车辆类型'] == '老总', '判定结果'] = '免费'
            
            writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
            
            # Sheet 1
            cols = ['车牌号码', '入场时间', '出场时间', '停车时长', '应收金额（元）', '优惠金额（元）', '实收金额（元）', '判定结果', '车辆类型']
            for c in cols:
                if c not in target_df.columns: target_df[c] = ''
            target_df[cols].sort_values('入场时间').to_excel(writer, sheet_name='最终判定结果', index=False)
            
            # Sheet 2
            grid_df = pd.DataFrame.from_dict(grid_data, orient='index')
            grid_df = grid_df.sort_index(axis=1) # 列排序
            # 行排序 (解决报错的关键点)
            if not grid_df.empty:
                grid_df.sort_index(level=1, inplace=True)
                grid_df.index = grid_df.index.get_level_values(0)
                
            grid_df.columns = [t.strftime('%Y-%m-%d %H:%M:%S') for t in grid_df.columns]
            grid_df.to_excel(writer, sheet_name='事件级校对表')
            
            writer.close()
            
            # ==========================
            # 样式美化
            # ==========================
            log_callback("正在应用颜色样式...")
            wb = load_workbook(output_file)
            
            # 美化 Sheet 1
            ws1 = wb['最终判定结果']
            red_font = Font(color="FF0000", bold=True)
            headers = [c.value for c in ws1[1]]
            if '车辆类型' in headers:
                t_idx = headers.index('车辆类型') + 1
                for row in ws1.iter_rows(min_row=2):
                    if row[t_idx-1].value == '老总':
                        for cell in row: cell.font = red_font
            
            # 美化 Sheet 2
            ws2 = wb['事件级校对表']
            green = PatternFill("solid", start_color="90EE90")
            yellow = PatternFill("solid", start_color="FFFF00")
            thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
            
            for row in ws2.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    val = cell.value
                    if val == '免': 
                        cell.fill = green
                        cell.border = thin_border
                    elif val == '收': 
                        cell.fill = yellow
                        cell.border = thin_border
            
            wb.save(output_file)
            progress_callback(100)
            log_callback(f"处理成功！结果已保存至:\n{output_file}")
            return True, "Success"

        except Exception as e:
            return False, str(e)


# ==========================================
# GUI 界面类
# ==========================================
class ParkingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("停车场收费智能判定系统 V7.0")
        self.root.geometry("650x750")
        self.root.resizable(False, False)
        
        self.logic = ParkingLogic()
        self.setup_ui()
        
    def setup_ui(self):
        # --- 样式配置 ---
        padding_opts = {'padx': 10, 'pady': 5}
        
        # 1. 文件选择区域
        file_frame = tk.LabelFrame(self.root, text="文件输入/输出", font=("微软雅黑", 10, "bold"))
        file_frame.pack(fill="x", **padding_opts)
        
        # 酒楼文件
        tk.Label(file_frame, text="酒楼原始记录:").grid(row=0, column=0, sticky="e")
        self.entry_hotel = tk.Entry(file_frame, width=50)
        self.entry_hotel.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="浏览", command=lambda: self.browse_file(self.entry_hotel)).grid(row=0, column=2)
        
        # 总表文件
        tk.Label(file_frame, text="停车场总记录:").grid(row=1, column=0, sticky="e")
        self.entry_total = tk.Entry(file_frame, width=50)
        self.entry_total.grid(row=1, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="浏览", command=lambda: self.browse_file(self.entry_total)).grid(row=1, column=2)
        
        # 输出路径
        tk.Label(file_frame, text="结果保存路径:").grid(row=2, column=0, sticky="e")
        self.entry_output = tk.Entry(file_frame, width=50)
        self.entry_output.grid(row=2, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="保存", command=lambda: self.save_file(self.entry_output)).grid(row=2, column=2)

        # 2. 参数设置区域
        param_frame = tk.LabelFrame(self.root, text="规则参数设置", font=("微软雅黑", 10, "bold"))
        param_frame.pack(fill="x", **padding_opts)
        
        # 基础免费车位
        tk.Label(param_frame, text="1. 基础免费车位(Base):").grid(row=0, column=0, sticky="e")
        self.entry_base = tk.Entry(param_frame, width=10)
        self.entry_base.insert(0, "8") # 默认值
        self.entry_base.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        tk.Label(param_frame, text="个 (例如：8)").grid(row=0, column=2, sticky="w")

        # 浮动车位 (手动输入)
        tk.Label(param_frame, text="2. 浮动车位(Floating):").grid(row=1, column=0, sticky="e")
        self.var_floating = tk.StringVar(value="0")  # 默认值

        # 【修改】: 去掉 state="readonly"，允许编辑
        self.entry_floating = tk.Entry(param_frame, width=10, textvariable=self.var_floating)

        self.entry_floating.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        # 【修改】: 提示文字改一下
        tk.Label(param_frame, text="个 (手动输入)").grid(row=1, column=2, sticky="w")

        # 老总车牌
        tk.Label(param_frame, text="3. 老总车牌(Boss):").grid(row=2, column=0, sticky="ne", pady=5)
        self.text_plates = tk.Text(param_frame, height=4, width=50)
        self.text_plates.grid(row=2, column=1, columnspan=2, padx=5, pady=5)
        self.text_plates.insert("1.0", "粤AB53314, 粤A8M00V")

        # 【修改】: 删掉下面这两行绑定代码，防止它自动覆盖你输入的值
        # self.text_plates.bind('<KeyRelease>', self.update_floating_count)
        # self.update_floating_count()
        
        tk.Label(param_frame, text="* 多个车牌请用逗号(,)分隔").grid(row=3, column=1, sticky="w", padx=5)

        # 3. 执行与进度
        run_frame = tk.Frame(self.root)
        run_frame.pack(fill="x", **padding_opts)
        
        self.btn_run = tk.Button(run_frame, text="开始计算", bg="#4CAF50", fg="white", font=("微软雅黑", 12, "bold"), height=2, command=self.start_process)
        self.btn_run.pack(fill="x", pady=5)
        
        self.progress = ttk.Progressbar(run_frame, orient="horizontal", length=600, mode="determinate")
        self.progress.pack(fill="x", pady=5)
        
        # 4. 日志输出
        log_frame = tk.LabelFrame(self.root, text="运行日志", font=("微软雅黑", 9))
        log_frame.pack(fill="both", expand=True, **padding_opts)
        
        self.text_log = tk.Text(log_frame, state="disabled", font=("Consolas", 9))
        self.text_log.pack(fill="both", expand=True)
        
    # --- 功能函数 ---
    
    def browse_file(self, entry_widget):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)
            
    def save_file(self, entry_widget):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)
            
    def update_floating_count(self, event=None):
        """根据输入的车牌自动计算浮动车位数"""
        content = self.text_plates.get("1.0", tk.END)
        # 简单的分割计算
        plates = [p for p in content.replace('，', ',').split(',') if p.strip()]
        count = len(plates)
        self.var_floating.set(str(count))
        
    def log(self, message):
        """向日志框写入信息 (线程安全)"""
        def _write():
            self.text_log.config(state="normal")
            self.text_log.insert(tk.END, message + "\n")
            self.text_log.see(tk.END)
            self.text_log.config(state="disabled")
        self.root.after(0, _write)
        
    def update_progress(self, value):
        """更新进度条 (线程安全)"""
        self.root.after(0, lambda: self.progress.configure(value=value))
        
    def toggle_inputs(self, enable=True):
        state = "normal" if enable else "disabled"
        self.btn_run.config(state=state)
        self.text_plates.config(state=state)
        
    def start_process(self):
        # 1. 获取输入
        f_hotel = self.entry_hotel.get().strip()
        f_total = self.entry_total.get().strip()
        f_out = self.entry_output.get().strip()
        base_quota_str = self.entry_base.get().strip()
        boss_plates = self.text_plates.get("1.0", tk.END).strip()

        # 【新增】: 获取浮动车位输入
        floating_quota_str = self.var_floating.get().strip()

        boss_plates = self.text_plates.get("1.0", tk.END).strip()

        # 2. 校验
        if not f_hotel or not os.path.exists(f_hotel):
            messagebox.showerror("错误", "请选择有效的酒楼原始记录文件！")
            return
        if not f_total or not os.path.exists(f_total):
            messagebox.showerror("错误", "请选择有效的停车场总记录文件！")
            return
        if not f_out:
            messagebox.showerror("错误", "请选择结果保存路径！")
            return
        if not base_quota_str.isdigit():
            messagebox.showerror("错误", "基础免费车位必须是整数！")
            return

        # 【新增】: 校验浮动车位
        if not floating_quota_str.isdigit():
            messagebox.showerror("错误", "浮动车位必须是整数！")
            return

        base_quota = int(base_quota_str)
        floating_quota = int(floating_quota_str)  # 转换类型
        
        # 3. 启动线程
        self.toggle_inputs(False)
        self.text_log.config(state="normal")
        self.text_log.delete("1.0", tk.END) # 清空日志
        self.text_log.config(state="disabled")
        self.progress['value'] = 0
        
        # threading.Thread(target=self.run_logic_thread, args=(f_hotel, f_total, f_out, base_quota, boss_plates)).start()
        # 【修改】: 传递 floating_quota 给线程函数
        threading.Thread(target=self.run_logic_thread,
                         args=(f_hotel, f_total, f_out, base_quota, floating_quota, boss_plates)).start()

    # 【修改后】：在参数列表中增加 floating_quota
    def run_logic_thread(self, f_hotel, f_total, f_out, base_quota, floating_quota, boss_plates):
        success, msg = self.logic.run(
            hotel_file=f_hotel,
            total_file=f_total,
            output_file=f_out,
            base_quota=base_quota,
            floating_quota=floating_quota,  # 别忘了传给逻辑类
            boss_plates_str=boss_plates,
            log_callback=self.log,
            progress_callback=self.update_progress
        )

        self.root.after(0, lambda: self.toggle_inputs(True))
        if success:
            self.root.after(0, lambda: messagebox.showinfo("完成", "计算完成！"))
        else:
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理过程中发生错误：\n{msg}"))

if __name__ == "__main__":
    root = tk.Tk()
    app = ParkingApp(root)
    root.mainloop()